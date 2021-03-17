#!/usr/bin/python3
# -*- coding: UTF-8 -*-
__Author__ = "Alvin Liu"

from openpyxl import load_workbook
import json


workbook = load_workbook('O365权限-20210317.xlsx')

# print(workbook.sheetnames)

detail_sheet = workbook['E1+E3']
detail_sheet_json = []
for i in range(2, detail_sheet.max_row + 1):
    sub_data = {
        "user_id": detail_sheet.cell(i, 6).value,
        "user_name": detail_sheet.cell(i, 2).value,
        "dept_name": "",
        "email": detail_sheet.cell(i, 6).value,
        "emp_id": "",
        "sys_id": "109"
    }
    detail_sheet_json.append(sub_data)
print('明细表的数据：', detail_sheet_json)

with open('./o365.json', "w+", encoding="utf-8") as  f:
    f.write(json.dumps(detail_sheet_json))
