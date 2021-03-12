#!/usr/bin/python3
# -*- coding: UTF-8 -*-
__Author__ = "Alvin Liu"


import json

from openpyxl import load_workbook
# workbook = xlrd.open_workbook_xls('202102养固健内销明细.xlsx')
workbook = load_workbook('202102养固健内销明细.xlsx')

# print(workbook.sheetnames)

detail_sheet = workbook['明细']
detail_sheet_json = []
for i in range(2, detail_sheet.max_row + 1):
    sub_data = (detail_sheet.cell(i, 2).value,
                str(detail_sheet.cell(i, 3).value),
                detail_sheet.cell(i, 4).value,
                detail_sheet.cell(i, 5).value,
                detail_sheet.cell(i, 6).value,
                detail_sheet.cell(i, 7).value,
                detail_sheet.cell(i, 8).value,
                detail_sheet.cell(i, 9).value,
                str(detail_sheet.cell(i, 10).value),
                str(detail_sheet.cell(i, 11).value))
    detail_sheet_json.append(sub_data)
print('明细表的数据：', detail_sheet_json)
# print(detail_sheet_json)


# address_sheet = workbook['配送地址']
# address_sheet_json = []
# for i in range(1, address_sheet.max_row + 1):
#     if address_sheet.cell(i, 1).value != '配送至':
#         sub_data = (address_sheet.cell(i, 1).value,
#                     address_sheet.cell(i, 2).value,
#                     address_sheet.cell(i, 3).value,
#                     str(address_sheet.cell(i, 4).value))
#         address_sheet_json.append(sub_data)
# print('配送地址的数据：', address_sheet_json)
# for i in address_sheet_json:
#     print('配送的地址项：', i)


# 执行msyql数据的更新
import pymysql
from popy.env import Env
employee_purchase = Env.getCfg().Config.EMPLOYEE_PURCHASE_SYSTEM


connection = pymysql.connect(employee_purchase['host'], employee_purchase['user'], employee_purchase['passwd'],
                             employee_purchase['db'])


# with connection.cursor() as cursor:
#     # Read a single record
#     # 查询需要更新的数据，左链接查出关联表的数据
#     # sql = """SELECT * from nx_order_address"""
#     # cursor.execute(sql)
#     # result = cursor.fetchall()
#     # 需要批量更新的sql
#     updatesql = """insert into nx_order_address(receive_office, detail_address, consignee_name, consignee_tel) values ((%s), (%s), (%s), (%s))"""
#     try:
#         execute_result = cursor.executemany(updatesql, address_sheet_json)
#         connection.commit()
#         # print('result:', result)
#         print('执行完成execute_result:', execute_result)
#
#     except Exception as e:
#         print("反正是写出问题了的：", e)
#     # finally:
#     #     connection.close()

with connection.cursor() as cursor:
    # Read a single record
    # 查询需要更新的数据，左链接查出关联表的数据
    # sql = """SELECT * from nx_order_address"""
    # cursor.execute(sql)
    # result = cursor.fetchall()
    # 需要批量更新的sql
    updatesql = """insert into nx_order_info(wechat_name, employee_id, employee_name, order_address, dept, order_number,
    order_title, order_desc, order_count, order_price) values ((%s), (%s), (%s), (%s), (%s), (%s), (%s), (%s), (%s), (%s))"""
    try:
        execute_result = cursor.executemany(updatesql, detail_sheet_json)
        connection.commit()
        # print('result:', result)
        print('执行完成execute_result:', execute_result)

    except Exception as e:
        print("反正是写出问题了的：", e)


# worksheets = workbook.sheet_names()
# print('worksheets is %s' %worksheets)
